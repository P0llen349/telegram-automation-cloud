#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TICKETS TO GPRS FORMATTER - FINAL STABLE VERSION
=================================================
This script transforms tickets CSV files to GPRS Excel format.
Version: Final - Corruption resistant
Author: Claude Assistant
Date: December 2025

FEATURES:
- Status mapping: 0→offline, -1→never online, 1→online
- Phone conversion: 962 to 079 format
- Coordinate processing with dummy detection
- Feedback subsheet for Problem/Solution data
- Robust error handling
"""

import pandas as pd
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path
import glob
import traceback
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
import pickle

class TicketsToGPRSFormatter:
    """
    Main class for converting tickets data to GPRS format
    """
    
    def __init__(self):
        """Initialize the formatter with folder paths"""
        try:
            # Use relative paths based on script location
            script_dir = Path(__file__).parent  # Ticket_Processing folder
            project_root = script_dir.parent  # Project_Organization folder

            self.raw_data_folder = str(project_root / "Tickets_Auto_Emails")
            self.transformed_data_folder = str(project_root / "Tickets_Auto_Emails" / "Transformed_DATA")

            # Create output folder if it doesn't exist
            os.makedirs(self.transformed_data_folder, exist_ok=True)

            print("Formatter initialized successfully")

        except Exception as e:
            print(f"Error initializing formatter: {e}")
            raise
    
    def load_latest_tickets(self):
        """
        Load the most recent CSV file from the raw data folder
        Returns: (dataframe, filepath)
        """
        try:
            # Find all CSV files
            csv_pattern = os.path.join(self.raw_data_folder, "*.csv")
            csv_files = glob.glob(csv_pattern)

            if not csv_files:
                raise FileNotFoundError(f"No CSV files found in {self.raw_data_folder}")

            # Get the latest file by modification time (most recently downloaded/modified)
            latest_file = max(csv_files, key=os.path.getmtime)
            filename = os.path.basename(latest_file)

            print(f"Loading latest file: {filename}")
            print(f"Total CSV files found: {len(csv_files)}")
            print(f"File modified: {datetime.fromtimestamp(os.path.getmtime(latest_file)).strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"\n[DEBUG] All CSV files found (sorted by date):")
            sorted_files = sorted(csv_files, key=os.path.getmtime, reverse=True)
            for idx, f in enumerate(sorted_files, 1):
                mtime = datetime.fromtimestamp(os.path.getmtime(f)).strftime("%Y-%m-%d %H:%M:%S")
                marker = " <-- SELECTED" if f == latest_file else ""
                print(f"  {idx}. {os.path.basename(f)} | {mtime}{marker}")
            print()
            
            # Try multiple encodings to handle Arabic text
            encodings = ['utf-8-sig', 'utf-8', 'cp1256', 'iso-8859-1']
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(latest_file, encoding=encoding)
                    
                    # Clean column names (remove extra spaces)
                    df.columns = [str(col).strip() for col in df.columns]
                    
                    print(f"Successfully loaded {len(df)} tickets using {encoding} encoding")
                    return df, latest_file
                    
                except Exception as encoding_error:
                    print(f"Failed with {encoding}: {str(encoding_error)[:100]}")
                    continue
            
            raise ValueError(f"Could not load file {filename} with any encoding")
            
        except Exception as e:
            print(f"Error loading tickets file: {e}")
            raise
    
    def convert_phone_number(self, phone):
        """
        Convert phone number from 962 format to 079 format
        Args: phone - phone number (any format)
        Returns: formatted phone number or empty string
        """
        try:
            if pd.isna(phone) or phone == "":
                return ""
            
            # Convert to string and clean
            phone_str = str(phone).replace('.0', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
            
            # Handle different formats
            if phone_str.startswith('962'):
                # Remove 962 country code
                remaining = phone_str[3:]
                if len(remaining) == 9 and remaining.startswith('7'):
                    return '0' + remaining
                    
            elif len(phone_str) == 9 and phone_str.startswith('7'):
                # Add 0 prefix
                return '0' + phone_str
                
            elif len(phone_str) == 10 and phone_str.startswith('07'):
                # Already in correct format
                return phone_str
            
            # Return as-is if can't convert
            return phone_str
            
        except Exception as e:
            print(f"Error converting phone {phone}: {e}")
            return str(phone) if not pd.isna(phone) else ""
    
    def process_online_status(self, status):
        """
        Process OnlineStatus values according to mapping rules
        Args: status - original status value
        Returns: mapped status string
        """
        try:
            if pd.isna(status):
                return ""
            
            status_str = str(status).strip()
            
            # Main mappings
            if status_str in ["0", "0.0"]:
                return "offline"
            elif status_str in ["-1", "-1.0"]:
                return "never online"
            elif status_str in ["1", "1.0"]:
                return "online"
            elif status_str in ["nan", ""]:
                return ""
            else:
                # Keep other values as-is
                return status_str
                
        except Exception as e:
            print(f"Error processing status {status}: {e}")
            return str(status) if not pd.isna(status) else ""
    
    def process_coordinates(self, lat_ticket, lon_ticket, lat_app, lon_app):
        """
        Process coordinates with dummy/no location detection
        Args: lat/lon from ticket and app sources
        Returns: (latitude, longitude) tuple
        """
        try:
            # Check for specific dummy coordinates in ticket columns FIRST
            if pd.notna(lat_ticket) and pd.notna(lon_ticket):
                try:
                    lat_ticket_val = float(lat_ticket)
                    lon_ticket_val = float(lon_ticket)
                    # Only flag as dummy if Latitude_ticket=30 AND Longitude_ticket=34
                    if lat_ticket_val == 30 and lon_ticket_val == 34:
                        return "dummy location", "dummy location"
                except (ValueError, TypeError):
                    pass
            
            lat_to_use = None
            lon_to_use = None
            
            # Try ticket coordinates first
            if pd.notna(lat_ticket) and pd.notna(lon_ticket):
                try:
                    lat_to_use = float(lat_ticket)
                    lon_to_use = float(lon_ticket)
                except (ValueError, TypeError):
                    pass
            
            # Fall back to app coordinates if needed
            if lat_to_use is None and pd.notna(lat_app) and pd.notna(lon_app):
                try:
                    lat_to_use = float(lat_app)
                    lon_to_use = float(lon_app)
                except (ValueError, TypeError):
                    pass
            
            # Check if we have valid coordinates
            if lat_to_use is not None and lon_to_use is not None:
                
                # Check for swapped coordinates (latitude should be smaller than longitude in Jordan)
                if lat_to_use > lon_to_use:
                    print(f"Warning: Swapped coordinates detected - Lat:{lat_to_use}, Lon:{lon_to_use} - Correcting...")
                    # Swap them to correct order
                    lat_to_use, lon_to_use = lon_to_use, lat_to_use
                
                # Return valid coordinates
                return round(lat_to_use, 6), round(lon_to_use, 6)
            
            # No valid coordinates found
            return "no location", "no location"
            
        except Exception as e:
            print(f"Error processing coordinates: {e}")
            return "no location", "no location"
    
    def create_street_building(self, category, street):
        """
        Create street_building field by combining category and street
        Args: category, street - input values
        Returns: combined string or empty
        """
        try:
            cat_str = ""
            street_str = ""
            
            if pd.notna(category) and str(category).strip() not in ["", "nan"]:
                cat_str = str(category).strip()
            
            if pd.notna(street) and str(street).strip() not in ["", "nan"]:
                street_str = str(street).strip()
            
            # Combine with dash if both exist
            if cat_str and street_str:
                return f"{cat_str}-{street_str}"
            elif cat_str:
                return cat_str
            elif street_str:
                return street_str
            else:
                return ""
                
        except Exception as e:
            print(f"Error creating street/building: {e}")
            return ""
    
    def transform_to_gprs_format(self, tickets_df):
        """
        Transform tickets DataFrame to GPRS format
        Args: tickets_df - input dataframe
        Returns: transformed dataframe
        """
        try:
            print("Transforming data to GPRS format...")
            
            result_df = pd.DataFrame()
            
            # 1. Serial number
            result_df['no'] = range(1, len(tickets_df) + 1)
            
            # 2. Status mapping
            if 'OnlineStatus' in tickets_df.columns:
                result_df['الحالة'] = tickets_df['OnlineStatus'].apply(self.process_online_status)
            else:
                result_df['الحالة'] = ""
            
            # 3. Basic fields
            result_df['رقم المرجع'] = tickets_df.get('Refcode', "")
            result_df['رقم العداد'] = tickets_df.get('Meter_no', "")
            
            # Process communication method and add "NO TECH" for empty meter numbers
            meter_numbers = tickets_df.get('Meter_no', pd.Series([""]*len(tickets_df)))
            comm_methods = tickets_df.get('Material_Group_Name', pd.Series([""]*len(tickets_df)))
            processed_comm_methods = []
            
            for meter, comm in zip(meter_numbers, comm_methods):
                if pd.isna(meter) or str(meter).strip() in ["", "nan", "0", "0.0"]:
                    processed_comm_methods.append("NO TECH")
                else:
                    processed_comm_methods.append(str(comm) if not pd.isna(comm) else "")
            
            result_df['طريقة الإتصال'] = processed_comm_methods
            result_df['اسم المشترك'] = tickets_df.get('customer_name', "")
            
            # 4. Phone conversion
            if 'phone' in tickets_df.columns:
                result_df['رقم المشترك'] = tickets_df['phone'].apply(self.convert_phone_number)
            else:
                result_df['رقم المشترك'] = ""
            
            # 5. Street/building combination
            category_col = tickets_df.get('Category', pd.Series([""]*len(tickets_df)))
            street_col = tickets_df.get('Street', pd.Series([""]*len(tickets_df)))
            
            street_building_list = []
            for cat, street in zip(category_col, street_col):
                street_building_list.append(self.create_street_building(cat, street))
            
            result_df['شارع_رقم بناية'] = street_building_list
            
            # 6. Office
            result_df['المكتب_المنطقة'] = tickets_df.get('OFFICE_NAME', "")
            
            # 7. Coordinates processing
            lat_coords = []
            lon_coords = []
            
            for i in range(len(tickets_df)):
                lat_ticket = tickets_df.get('Latitude_Ticket', pd.Series([None]*len(tickets_df))).iloc[i]
                lon_ticket = tickets_df.get('Longitude_Ticket', pd.Series([None]*len(tickets_df))).iloc[i]
                lat_app = tickets_df.get('Latitude_App', pd.Series([None]*len(tickets_df))).iloc[i]
                lon_app = tickets_df.get('Longitude_app', pd.Series([None]*len(tickets_df))).iloc[i]
                
                lat, lon = self.process_coordinates(lat_ticket, lon_ticket, lat_app, lon_app)
                lat_coords.append(lat)
                lon_coords.append(lon)
            
            result_df['Latitude'] = lat_coords
            result_df['Longitude'] = lon_coords
            
            # 8. Submit date
            result_df['SubmitDate'] = tickets_df.get('SubmitDate', "")
            
            print(f"Successfully created {len(result_df)} GPRS records")
            return result_df
            
        except Exception as e:
            print(f"Error transforming data: {e}")
            print(traceback.format_exc())
            raise
    
    def create_feedback_sheet(self, tickets_df, main_gprs_df):
        """
        Create feedback subsheet for tickets with Problem/Solution data
        AND remove these entries from the main sheet
        Args: tickets_df - original data, main_gprs_df - transformed main data
        Returns: (feedback dataframe, updated main dataframe) or (None, main_gprs_df)
        """
        try:
            print("Creating feedback subsheet...")
            
            # Check if Problem/Solution columns exist
            if 'Solution' not in tickets_df.columns and 'Problem' not in tickets_df.columns:
                print("No Problem/Solution columns found")
                return None, main_gprs_df
            
            # Filter rows with Solution OR Problem data
            solution_mask = pd.Series([False] * len(tickets_df))
            problem_mask = pd.Series([False] * len(tickets_df))
            
            if 'Solution' in tickets_df.columns:
                solution_mask = (tickets_df['Solution'].notna() & 
                               (tickets_df['Solution'].astype(str).str.strip() != '') & 
                               (tickets_df['Solution'].astype(str) != 'nan'))
            
            if 'Problem' in tickets_df.columns:
                problem_mask = (tickets_df['Problem'].notna() & 
                              (tickets_df['Problem'].astype(str).str.strip() != '') & 
                              (tickets_df['Problem'].astype(str) != 'nan'))
            
            feedback_mask = solution_mask | problem_mask
            
            if not feedback_mask.any():
                print("No feedback data found (empty Problem/Solution fields)")
                return None, main_gprs_df
            
            # Get corresponding rows from main GPRS sheet for feedback
            feedback_indices = tickets_df.index[feedback_mask].tolist()
            feedback_final = main_gprs_df.loc[main_gprs_df.index.isin(feedback_indices)].copy()
            
            # Reset serial numbers for feedback sheet
            feedback_final['no'] = range(1, len(feedback_final) + 1)
            
            # Add Problem and Solution columns
            feedback_tickets = tickets_df.loc[feedback_indices]
            feedback_final['Problem'] = feedback_tickets.get('Problem', "").values
            feedback_final['Solution'] = feedback_tickets.get('Solution', "").values
            
            # Remove feedback entries from main sheet
            updated_main_gprs = main_gprs_df.loc[~main_gprs_df.index.isin(feedback_indices)].copy()
            
            # Reset serial numbers for updated main sheet
            updated_main_gprs['no'] = range(1, len(updated_main_gprs) + 1)
            
            print(f"Feedback sheet created with {len(feedback_final)} records")
            print(f"Main sheet updated: {len(main_gprs_df)} -> {len(updated_main_gprs)} records")
            return feedback_final, updated_main_gprs
            
        except Exception as e:
            print(f"Error creating feedback sheet: {e}")
            print(traceback.format_exc())
            return None, main_gprs_df
    
    def create_daily_ticket_analysis(self, main_df):
        """
        Create pivot table analyzing tickets by connection type and day
        Args: main_df - main GPRS data
        Returns: pivot table dataframe or None
        """
        try:
            print("Creating daily ticket count analysis...")
            
            # Check if required columns exist
            if 'طريقة الإتصال' not in main_df.columns or 'SubmitDate' not in main_df.columns:
                print("Missing required columns for analysis")
                return None
            
            # Create a working copy
            analysis_df = main_df.copy()
            
            # Parse dates and extract date only (without time)
            parsed_dates = []
            valid_rows = []
            
            for idx, submit_date_str in enumerate(analysis_df['SubmitDate']):
                submit_date = self.parse_submit_date(submit_date_str)
                if submit_date is not None:
                    # Extract just the date part (YYYY-MM-DD)
                    date_only = submit_date.strftime('%Y-%m-%d')
                    parsed_dates.append(date_only)
                    valid_rows.append(idx)
                
            if not parsed_dates:
                print("No valid dates found for analysis")
                return None
            
            # Filter to only valid date rows
            analysis_df = analysis_df.iloc[valid_rows].copy()
            analysis_df['Date'] = parsed_dates
            
            # Clean connection type data
            analysis_df['طريقة الإتصال'] = analysis_df['طريقة الإتصال'].fillna('Unknown')
            analysis_df['طريقة الإتصال'] = analysis_df['طريقة الإتصال'].astype(str).str.strip()
            
            # Create full pivot table for charts (with all dates)
            full_pivot_table = pd.pivot_table(
                analysis_df,
                values='no',  # Count tickets using the serial number
                index='طريقة الإتصال',
                columns='Date',
                aggfunc='count',
                fill_value=0,
                margins=True,  # Add totals
                margins_name='Total'
            )
            
            # Sort dates in chronological order (excluding Total column)
            date_columns = [col for col in full_pivot_table.columns if col != 'Total']
            date_columns_sorted = sorted(date_columns)
            
            # Reorder full table columns: sorted dates + Total
            if 'Total' in full_pivot_table.columns:
                final_columns = date_columns_sorted + ['Total']
                full_pivot_table = full_pivot_table[final_columns]
            
            # Sort connection types by total count (descending)
            if 'Total' in full_pivot_table.index:
                # Remove Total row temporarily for sorting
                total_row = full_pivot_table.loc['Total']
                full_pivot_table = full_pivot_table.drop('Total')
                
                # Sort by Total column (descending)
                if 'Total' in full_pivot_table.columns:
                    full_pivot_table = full_pivot_table.sort_values('Total', ascending=False)
                
                # Add Total row back at the bottom
                full_pivot_table.loc['Total'] = total_row
            
            # Create simplified pivot table with just totals
            simple_pivot_table = full_pivot_table[['Total']].copy()
            
            # Store both tables for different uses
            result = {
                'full_table': full_pivot_table,
                'simple_table': simple_pivot_table,
                'date_columns': date_columns_sorted
            }
            
            print(f"Created pivot tables with {len(full_pivot_table)-1} connection types across {len(date_columns_sorted)} days")
            return result
            
        except Exception as e:
            print(f"Error creating daily analysis: {e}")
            print(traceback.format_exc())
            return None
    
    def create_charts_in_workbook(self, workbook, sheet_name, pivot_data, main_df):
        """
        Create simple pie chart for communication technology distribution
        Args: workbook - Excel workbook, sheet_name - sheet name, pivot_data - pivot table data, main_df - main dataframe
        """
        try:
            if pivot_data is None or 'simple_table' not in pivot_data:
                return
            
            print("Creating communication technology pie chart...")
            worksheet = workbook[sheet_name]
            simple_table = pivot_data['simple_table']
            
            # Get data dimensions 
            connection_types = len(simple_table) - 1  # Exclude Total row if exists
            
            if connection_types <= 0:
                print("Not enough data for charts")
                return
            
            from openpyxl.chart import PieChart, Reference
            from openpyxl.chart.label import DataLabelList
            
            # Pie Chart for communication technology distribution
            pie_chart = PieChart()
            pie_chart.title = "Tickets by Communication Technology"
            pie_chart.width = 15
            pie_chart.height = 12
            
            # Data range: Total column (should be column B)
            pie_data = Reference(worksheet, min_col=2, min_row=2, max_col=2, max_row=connection_types + 1)
            pie_labels = Reference(worksheet, min_col=1, min_row=2, max_col=1, max_row=connection_types + 1)
            
            pie_chart.add_data(pie_data)
            pie_chart.set_categories(pie_labels)
            
            # Add data labels
            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showPercent = True
            pie_chart.dataLabels.showVal = True
            pie_chart.dataLabels.showCatName = True
            
            # Position pie chart
            worksheet.add_chart(pie_chart, "A8")
            
            print("Created communication technology pie chart")
            
        except Exception as e:
            print(f"Error creating charts: {e}")
            print(traceback.format_exc())
    
    def create_tech_status_pivot_table(self, main_df):
        """
        Create pivot table for technology-status combinations
        Args: main_df - main dataframe
        Returns: pivot table dataframe
        """
        try:
            print("Creating technology-status pivot table...")
            
            # Check if required columns exist
            if 'طريقة الإتصال' not in main_df.columns or 'الحالة' not in main_df.columns:
                print("Required columns not found in dataframe")
                return None
            
            # Create a working copy
            analysis_df = main_df.copy()
            
            # Clean the data
            analysis_df['طريقة الإتصال'] = analysis_df['طريقة الإتصال'].fillna('Unknown')
            analysis_df['الحالة'] = analysis_df['الحالة'].fillna('blank')
            
            # Convert to string and clean
            analysis_df['طريقة الإتصال'] = analysis_df['طريقة الإتصال'].astype(str).str.strip()
            analysis_df['الحالة'] = analysis_df['الحالة'].astype(str).str.strip()
            
            # Replace empty or nan values
            analysis_df.loc[analysis_df['الحالة'].isin(['', 'nan', 'None', 'NaN']), 'الحالة'] = 'blank'
            analysis_df.loc[analysis_df['طريقة الإتصال'].isin(['', 'nan', 'None', 'NaN']), 'طريقة الإتصال'] = 'Unknown'
            
            # Create pivot table
            pivot_table = pd.pivot_table(
                analysis_df,
                values='no',  # Count tickets using the serial number
                index='طريقة الإتصال',
                columns='الحالة',
                aggfunc='count',
                fill_value=0
            )
            
            print(f"Pivot table shape: {pivot_table.shape}")
            print(f"Technologies: {list(pivot_table.index)}")
            print(f"Statuses: {list(pivot_table.columns)}")
            
            return pivot_table
            
        except Exception as e:
            print(f"Error creating tech-status pivot table: {e}")
            print(traceback.format_exc())
            return None

    def create_combined_tech_status_chart(self, workbook, sheet_name, main_df):
        """
        Create pie chart showing communication technology and status combinations using a separate data sheet
        Args: workbook - Excel workbook, sheet_name - sheet name, main_df - main dataframe
        """
        try:
            print("Creating combined technology-status pie chart...")
            
            # Create pivot table first
            pivot_table = self.create_tech_status_pivot_table(main_df)
            if pivot_table is None:
                return
            
            # Create a separate sheet for chart data
            try:
                data_sheet = workbook.create_sheet("ChartData")
            except:
                data_sheet = workbook["ChartData"]
            
            # Write headers
            data_sheet['A1'] = 'Technology-Status'
            data_sheet['B1'] = 'Count'
            
            # Flatten pivot table into combinations
            row_idx = 2
            for tech_index, tech_row in pivot_table.iterrows():
                for status_col, count in tech_row.items():
                    if count > 0:  # Only include non-zero combinations
                        combination = f"{tech_index}-{status_col}"
                        data_sheet[f'A{row_idx}'] = combination
                        data_sheet[f'B{row_idx}'] = count
                        row_idx += 1
            
            print(f"Wrote chart data to separate sheet, rows 2 to {row_idx-1}")
            
            # Create pie chart on the main sheet
            from openpyxl.chart import PieChart, Reference
            from openpyxl.chart.label import DataLabelList
            
            combo_pie_chart = PieChart()
            combo_pie_chart.title = "Technology-Status Distribution"
            combo_pie_chart.width = 15
            combo_pie_chart.height = 12
            
            # Data range from the separate sheet
            if row_idx > 2:  # Make sure we have data
                data_range = Reference(data_sheet, min_col=2, min_row=2, max_col=2, max_row=row_idx-1)
                labels_range = Reference(data_sheet, min_col=1, min_row=2, max_col=1, max_row=row_idx-1)
                
                combo_pie_chart.add_data(data_range)
                combo_pie_chart.set_categories(labels_range)
                
                # Add data labels
                combo_pie_chart.dataLabels = DataLabelList()
                combo_pie_chart.dataLabels.showPercent = True
                combo_pie_chart.dataLabels.showVal = True
                combo_pie_chart.dataLabels.showCatName = True
                
                # Add chart to the main sheet
                main_sheet = workbook[sheet_name]
                main_sheet.add_chart(combo_pie_chart, "Q8")
                
                print(f"Created combined chart with {row_idx - 2} technology-status combinations")
            else:
                print("No data available for combined chart")
            
        except Exception as e:
            print(f"Error creating combined chart: {e}")
            print(traceback.format_exc())
    
    def create_location_charts(self, workbook, sheet_name, location_data):
        """
        Create charts for location analysis data
        Args: workbook - Excel workbook, sheet_name - sheet name, location_data - location pivot data
        """
        try:
            if location_data is None or len(location_data) == 0:
                return
                
            print("Creating location analysis charts...")
            worksheet = workbook[sheet_name]
            
            from openpyxl.chart import PieChart, BarChart, Reference
            from openpyxl.chart.label import DataLabelList
            
            # Pie Chart for location distribution
            pie_chart = PieChart()
            pie_chart.title = "Location Data Distribution"
            pie_chart.width = 15
            pie_chart.height = 12
            
            # Data range: Count column
            pie_data = Reference(worksheet, min_col=2, min_row=2, max_col=2, max_row=len(location_data) + 1)
            pie_labels = Reference(worksheet, min_col=1, min_row=2, max_col=1, max_row=len(location_data) + 1)
            
            pie_chart.add_data(pie_data)
            pie_chart.set_categories(pie_labels)
            
            # Add data labels
            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showPercent = True
            pie_chart.dataLabels.showVal = True
            pie_chart.dataLabels.showCatName = True
            
            # Position pie chart
            worksheet.add_chart(pie_chart, "A8")
            
            # Bar Chart for location counts
            bar_chart = BarChart()
            bar_chart.title = "Location Data Quality Analysis"
            bar_chart.type = "col"
            bar_chart.style = 2
            bar_chart.y_axis.title = 'Number of Tickets'
            bar_chart.x_axis.title = 'Location Type'
            bar_chart.width = 15
            bar_chart.height = 10
            
            # Add data
            bar_chart.add_data(pie_data, titles_from_data=False)
            bar_chart.set_categories(pie_labels)
            
            # Position bar chart
            worksheet.add_chart(bar_chart, "A25")
            
            print("Created location analysis charts: Pie and Bar")
            
        except Exception as e:
            print(f"Error creating location charts: {e}")
            print(traceback.format_exc())
    
    def parse_submit_date(self, date_str):
        """
        Parse SubmitDate string to datetime object
        Handles multiple date formats
        Args: date_str - date string from SubmitDate column
        Returns: datetime object or None
        """
        if pd.isna(date_str) or date_str == "":
            return None
        
        try:
            date_str = str(date_str).strip()
            
            # Common formats to try
            formats = [
                "%Y-%m-%d %H:%M:%S.%f",  # 2025-12-20 14:30:45.000
                "%Y-%m-%d %H:%M:%S",     # 2025-12-20 14:30:45
                "%Y-%m-%d",              # 2025-12-20
                "%d-%m-%Y %H:%M:%S",     # 20-12-2025 14:30:45
                "%d-%m-%Y",              # 20-12-2025
                "%b %d %Y %I:%M%p",      # Dec 20 2025 2:30PM
                "%b  %d %Y %I:%M%p",     # Dec  9 2025 2:30PM (double space)
            ]
            
            for fmt in formats:
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue
            
            # If none work, try pandas to_datetime as fallback
            return pd.to_datetime(date_str, errors='coerce')
            
        except Exception as e:
            return None
    
    def apply_date_highlighting(self, output_path, main_df, feedback_df=None):
        """
        Apply date-based row highlighting to Excel file
        - Red: Tickets older than 1 week
        - Light Green: Tickets from today (same date as sheet)
        Args: output_path - Excel file path, main_df - main data, feedback_df - feedback data
        """
        try:
            print("Applying date-based highlighting...")
            
            # Load the workbook
            workbook = load_workbook(output_path)
            
            # Get today's date (sheet creation date)
            today = datetime.now().date()
            yesterday = today - timedelta(days=1)
            one_week_ago = today - timedelta(days=7)
            
            # Define highlight colors
            red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")      # Light red
            green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")   # Light green
            
            # Process main sheet
            if 'GPRS_Data' in workbook.sheetnames:
                worksheet = workbook['GPRS_Data']
                self._highlight_sheet_rows(worksheet, main_df, today, yesterday, one_week_ago, red_fill, green_fill)
            
            # Process feedback sheet
            if feedback_df is not None and 'feedback' in workbook.sheetnames:
                worksheet = workbook['feedback']
                self._highlight_sheet_rows(worksheet, feedback_df, today, yesterday, one_week_ago, red_fill, green_fill)
            
            # Save the workbook
            workbook.save(output_path)
            print("Date highlighting applied successfully")
            
        except Exception as e:
            print(f"Error applying date highlighting: {e}")
            print(traceback.format_exc())
    
    def _highlight_sheet_rows(self, worksheet, df, today, yesterday, one_week_ago, red_fill, green_fill):
        """
        Apply highlighting to a specific worksheet
        Args: worksheet - Excel worksheet, df - dataframe, dates and fills
        """
        try:
            # Find SubmitDate column
            submit_date_col = None
            for col_idx, cell in enumerate(worksheet[1], 1):  # Header row
                if cell.value == 'SubmitDate':
                    submit_date_col = col_idx
                    break
            
            if submit_date_col is None:
                print("SubmitDate column not found in sheet")
                return
            
            red_count = 0
            green_count = 0
            
            # Process each data row (starting from row 2, skipping header)
            for row_idx in range(2, len(df) + 2):
                try:
                    # Get the SubmitDate value from dataframe
                    df_row_idx = row_idx - 2  # Convert to 0-based index
                    if df_row_idx >= len(df):
                        continue
                    
                    submit_date_str = df.iloc[df_row_idx]['SubmitDate']
                    submit_date = self.parse_submit_date(submit_date_str)
                    
                    if submit_date is None:
                        continue
                    
                    submit_date_only = submit_date.date()
                    
                    # Determine highlight color
                    fill_color = None
                    
                    # Check if older than 1 week (priority - red)
                    if submit_date_only <= one_week_ago:
                        fill_color = red_fill
                        red_count += 1
                    
                    # Check if 1 day old (yesterday's entries - green)
                    elif submit_date_only == yesterday:
                        fill_color = green_fill
                        green_count += 1
                    
                    # Apply highlighting to entire row
                    if fill_color is not None:
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col)
                            cell.fill = fill_color
                
                except Exception as row_error:
                    # Skip problematic rows
                    continue
            
            print(f"  Highlighted {red_count} old rows (>1 week) in red")
            print(f"  Highlighted {green_count} new rows (1 day old) in green")
            
        except Exception as e:
            print(f"Error highlighting sheet: {e}")
    
    def create_location_analysis(self, main_df):
        """
        Create pivot table analyzing tickets by location type (real, dummy, no location)
        Args: main_df - main GPRS data
        Returns: location analysis dataframe or None
        """
        try:
            print("Creating location analysis pivot table...")
            
            # Check if Latitude column exists
            if 'Latitude' not in main_df.columns:
                print("Missing Latitude column for location analysis")
                return None
            
            # Create location categories
            analysis_df = main_df.copy()
            location_categories = []
            
            for lat in analysis_df['Latitude']:
                if lat == 'no location':
                    location_categories.append('No Location')
                elif lat == 'dummy location':
                    location_categories.append('Dummy Location')
                else:
                    location_categories.append('Real Location')
            
            analysis_df['Location_Type'] = location_categories
            
            # Count by location type
            location_counts = analysis_df['Location_Type'].value_counts().to_frame()
            location_counts.columns = ['Count']
            
            # Calculate percentages
            total_count = len(analysis_df)
            location_counts['Percentage'] = (location_counts['Count'] / total_count * 100).round(1)
            
            # Sort by count (descending)
            location_counts = location_counts.sort_values('Count', ascending=False)
            
            print(f"Location analysis created with {len(location_counts)} location types")
            return location_counts
            
        except Exception as e:
            print(f"Error creating location analysis: {e}")
            print(traceback.format_exc())
            return None
    
    def save_excel_file(self, main_df, feedback_df=None, pivot_df=None, original_df=None):
        """
        Save data to Excel file with proper formatting and date highlighting
        Args: main_df - main data, feedback_df - feedback data (optional), pivot_df - pivot table (optional), original_df - original CSV data (optional)
        Returns: output file path
        """
        try:
            # Generate output filename
            timestamp = datetime.now().strftime("%d-%m-%Y")
            output_filename = f"{timestamp}_GPRS_CHARTS_v15.xlsx"
            output_path = os.path.join(self.transformed_data_folder, output_filename)

            print(f"Saving to: {output_path}")

            # Sort main dataframe by communication technology
            if 'طريقة الإتصال' in main_df.columns:
                main_df = main_df.sort_values('طريقة الإتصال', na_position='last')
                # Reset serial numbers after sorting
                main_df['no'] = range(1, len(main_df) + 1)
                print(f"Sorted main sheet by communication technology")

            # Sort feedback dataframe by communication technology
            if feedback_df is not None and 'طريقة الإتصال' in feedback_df.columns:
                feedback_df = feedback_df.sort_values('طريقة الإتصال', na_position='last')
                # Reset serial numbers after sorting
                feedback_df['no'] = range(1, len(feedback_df) + 1)
                print(f"Sorted feedback sheet by communication technology")

            # Save with Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

                # Main sheet
                main_df.to_excel(writer, index=False, sheet_name='GPRS_Data')
                
                # Feedback sheet if available
                if feedback_df is not None:
                    feedback_df.to_excel(writer, index=False, sheet_name='feedback')
                    print(f"Added feedback sheet with {len(feedback_df)} records")
                
                # Pivot table sheet if available
                if pivot_df is not None:
                    # Save full table for charts, but we'll replace it with simple table later
                    full_table = pivot_df['full_table'] 
                    full_table.to_excel(writer, index=True, sheet_name='Daily ticket count')
                    print(f"Added daily analysis sheet with {len(full_table)} connection types")
                
                # Location analysis pivot table
                location_pivot = self.create_location_analysis(main_df)
                if location_pivot is not None:
                    location_pivot.to_excel(writer, index=True, sheet_name='Location Analysis')
                    print(f"Added location analysis sheet with {len(location_pivot)} location types")

                # Original data sheet (at the end)
                if original_df is not None:
                    original_df.to_excel(writer, index=False, sheet_name='original Data')
                    print(f"Added original data sheet with {len(original_df)} records")

                # Auto-adjust column widths for readability
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                            except:
                                pass
                        
                        # Set column width (max 30 characters)
                        adjusted_width = min(max_length + 2, 30)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply date-based highlighting after saving
            self.apply_date_highlighting(output_path, main_df, feedback_df)
            
            # Add charts to pivot table sheet if pivot data is available
            if pivot_df is not None:
                print("Adding charts and simplifying pivot table...")
                workbook = load_workbook(output_path)
                
                # Create charts using full data (pass the updated main_df)
                self.create_charts_in_workbook(workbook, 'Daily ticket count', pivot_df, main_df)
                
                # Add the second chart BEFORE clearing the worksheet
                self.create_combined_tech_status_chart(workbook, 'Daily ticket count', main_df)
                
                # Replace full table with simple table (just totals)
                worksheet = workbook['Daily ticket count']
                
                # Clear the worksheet first
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.value = None
                
                # Add simplified table at the top
                simple_table = pivot_df['simple_table']
                
                # Write headers
                worksheet['A1'] = 'طريقة الإتصال'
                worksheet['B1'] = 'Total'
                
                # Write data
                for idx, (index, row) in enumerate(simple_table.iterrows(), 2):
                    worksheet[f'A{idx}'] = index
                    worksheet[f'B{idx}'] = row['Total']
                
                # Add location charts if location data is available
                location_pivot = self.create_location_analysis(main_df)
                if location_pivot is not None:
                    self.create_location_charts(workbook, 'Location Analysis', location_pivot)
                
                workbook.save(output_path)
                print("Charts added and pivot table simplified successfully")
            
            print(f"File saved successfully: {output_filename}")
            return output_path
            
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            print(traceback.format_exc())
            raise
    
    def generate_summary(self, main_df, feedback_df=None):
        """
        Generate summary statistics for the transformation
        Args: main_df - main data, feedback_df - feedback data (optional)
        """
        try:
            print("\n" + "="*60)
            print("           TRANSFORMATION SUMMARY")
            print("="*60)
            
            # Basic stats
            print(f"Total records processed: {len(main_df)}")
            
            # Date analysis for highlighting
            if 'SubmitDate' in main_df.columns:
                today = datetime.now().date()
                yesterday = today - timedelta(days=1)
                one_week_ago = today - timedelta(days=7)
                
                old_count = 0
                yesterday_count = 0
                parsed_dates = 0
                
                for submit_date_str in main_df['SubmitDate']:
                    submit_date = self.parse_submit_date(submit_date_str)
                    if submit_date is not None:
                        parsed_dates += 1
                        submit_date_only = submit_date.date()
                        
                        if submit_date_only <= one_week_ago:
                            old_count += 1
                        elif submit_date_only == yesterday:
                            yesterday_count += 1
                
                print(f"\nDate-based highlighting:")
                print(f"  Red (>1 week old): {old_count} tickets")
                print(f"  Green (1 day old): {yesterday_count} tickets")
                print(f"  Total with valid dates: {parsed_dates}/{len(main_df)}")
            
            # Status distribution
            if 'الحالة' in main_df.columns:
                status_counts = main_df['الحالة'].value_counts()
                print(f"\nStatus distribution:")
                for status, count in status_counts.head(5).items():
                    pct = (count / len(main_df)) * 100
                    print(f"  {status}: {count} ({pct:.1f}%)")
            
            # Phone conversion stats
            if 'رقم المشترك' in main_df.columns:
                phone_079 = main_df['رقم المشترك'].str.startswith('07', na=False).sum()
                print(f"\nPhone conversion:")
                print(f"  079 format: {phone_079}/{len(main_df)} ({(phone_079/len(main_df)*100):.1f}%)")
            
            # Coordinate stats
            if 'Latitude' in main_df.columns:
                no_location = (main_df['Latitude'] == 'no location').sum()
                dummy_location = (main_df['Latitude'] == 'dummy location').sum()
                real_coords = len(main_df) - no_location - dummy_location
                
                print(f"\nCoordinate distribution:")
                print(f"  Real coordinates: {real_coords} ({(real_coords/len(main_df)*100):.1f}%)")
                print(f"  Dummy locations: {dummy_location} ({(dummy_location/len(main_df)*100):.1f}%)")
                print(f"  No location: {no_location} ({(no_location/len(main_df)*100):.1f}%)")
            
            # Feedback stats
            if feedback_df is not None:
                print(f"\nFeedback records: {len(feedback_df)}")
            
            print("="*60)
            
        except Exception as e:
            print(f"Error generating summary: {e}")

    def get_oauth_credentials(self):
        """Get OAuth credentials for user's Google account"""
        script_dir = Path(__file__).parent
        project_root = script_dir.parent

        SCOPES = ['https://www.googleapis.com/auth/drive']
        creds = None

        # Token file stores user's access and refresh tokens
        token_path = project_root / 'token.pickle'
        oauth_creds_path = project_root / 'oauth_credentials.json'

        # Check if we have valid credentials
        if token_path.exists():
            with open(token_path, 'rb') as token:
                creds = pickle.load(token)

        # If no valid credentials, let user log in
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                print("Refreshing access token...")
                creds.refresh(Request())
            else:
                if not oauth_creds_path.exists():
                    print(f"ERROR: OAuth credentials not found at {oauth_creds_path}")
                    print("Please download oauth_credentials.json from Google Cloud Console")
                    return None

                print("Opening browser for Google authentication...")
                flow = InstalledAppFlow.from_client_secrets_file(
                    str(oauth_creds_path), SCOPES)
                creds = flow.run_local_server(port=8080)

            # Save credentials for next time
            with open(token_path, 'wb') as token:
                pickle.dump(creds, token)

        return creds

    def import_to_permanent_sheet(self, temp_spreadsheet_id, credentials):
        """
        Import all sheets from temporary spreadsheet to permanent spreadsheet
        Args: temp_spreadsheet_id - ID of temporary Google Sheet, credentials - OAuth credentials
        Returns: permanent sheet URL or None
        """
        try:
            print("\n" + "="*60)
            print("      IMPORTING TO PERMANENT GOOGLE SHEET")
            print("="*60)
            print("")

            # Your permanent Google Sheet ID
            PERMANENT_SHEET_ID = "13x58yfkrvA9_7bo-Wtzw6EwcPVCjE8x2IUmkF8c6Aro"

            # Build Sheets API service
            sheets_service = build('sheets', 'v4', credentials=credentials)
            drive_service = build('drive', 'v3', credentials=credentials)

            # ========================================================
            # STEP 1: Get all sheet names from temporary spreadsheet
            # ========================================================
            print("[STEP 1] Reading sheets from temporary Google Sheet...")

            temp_spreadsheet = sheets_service.spreadsheets().get(
                spreadsheetId=temp_spreadsheet_id
            ).execute()

            temp_sheets = temp_spreadsheet.get('sheets', [])
            temp_sheet_names = [sheet['properties']['title'] for sheet in temp_sheets]

            print(f"   [OK] Found {len(temp_sheet_names)} sheets to import:")
            for name in temp_sheet_names:
                print(f"      - {name}")
            print("")

            # ========================================================
            # STEP 2: Get existing sheets from permanent spreadsheet
            # ========================================================
            print("[STEP 2] Reading existing sheets from permanent Google Sheet...")

            permanent_spreadsheet = sheets_service.spreadsheets().get(
                spreadsheetId=PERMANENT_SHEET_ID
            ).execute()

            permanent_sheets = permanent_spreadsheet.get('sheets', [])

            print(f"   [OK] Found {len(permanent_sheets)} existing sheet(s)")
            print("")

            # ========================================================
            # STEP 3: Delete all existing sheets from permanent
            # ========================================================
            print("[STEP 3] Deleting old sheets from permanent Google Sheet...")

            requests = []

            # Delete all sheets except the last one (can't delete all sheets)
            for i, sheet in enumerate(permanent_sheets):
                sheet_id = sheet['properties']['sheetId']
                sheet_title = sheet['properties']['title']

                # Keep the last sheet temporarily
                if i < len(permanent_sheets) - 1:
                    requests.append({
                        'deleteSheet': {
                            'sheetId': sheet_id
                        }
                    })
                    print(f"   [DELETE] Queued for deletion: {sheet_title}")

            # Execute deletions
            if requests:
                sheets_service.spreadsheets().batchUpdate(
                    spreadsheetId=PERMANENT_SHEET_ID,
                    body={'requests': requests}
                ).execute()
                print(f"   [OK] Deleted {len(requests)} old sheet(s)")

            print("")

            # ========================================================
            # STEP 4: Copy all sheets from temp to permanent
            # ========================================================
            print("[STEP 4] Copying new sheets to permanent Google Sheet...")

            for sheet in temp_sheets:
                sheet_id = sheet['properties']['sheetId']
                sheet_title = sheet['properties']['title']

                # Copy sheet from temp to permanent
                copy_request = sheets_service.spreadsheets().sheets().copyTo(
                    spreadsheetId=temp_spreadsheet_id,
                    sheetId=sheet_id,
                    body={'destinationSpreadsheetId': PERMANENT_SHEET_ID}
                ).execute()

                print(f"   [OK] Copied: {sheet_title}")

            print("")

            # ========================================================
            # STEP 5: Delete the last remaining old sheet
            # ========================================================
            print("[STEP 5] Cleaning up remaining old sheet...")

            # Get updated permanent spreadsheet
            updated_permanent = sheets_service.spreadsheets().get(
                spreadsheetId=PERMANENT_SHEET_ID
            ).execute()

            updated_sheets = updated_permanent.get('sheets', [])

            # Find and delete the old sheet (not in temp_sheet_names)
            for sheet in updated_sheets:
                sheet_id = sheet['properties']['sheetId']
                sheet_title = sheet['properties']['title']

                if sheet_title not in temp_sheet_names:
                    sheets_service.spreadsheets().batchUpdate(
                        spreadsheetId=PERMANENT_SHEET_ID,
                        body={'requests': [{'deleteSheet': {'sheetId': sheet_id}}]}
                    ).execute()
                    print(f"   [OK] Deleted old sheet: {sheet_title}")
                    break

            print("")

            # ========================================================
            # STEP 6: Delete temporary Google Sheet
            # ========================================================
            print("[STEP 6] Deleting temporary Google Sheet (keeping Drive clean)...")

            try:
                drive_service.files().delete(fileId=temp_spreadsheet_id).execute()
                print(f"   [OK] Temporary Google Sheet deleted")
            except Exception as delete_error:
                print(f"   [WARNING] Could not delete temp file: {delete_error}")

            print("")

            # ========================================================
            # SUCCESS!
            # ========================================================
            permanent_url = f"https://docs.google.com/spreadsheets/d/{PERMANENT_SHEET_ID}/edit"

            print("="*60)
            print("[SUCCESS] Data imported to permanent Google Sheet!")
            print("="*60)
            print("")
            print(f"  Imported Sheets:")
            for name in temp_sheet_names:
                print(f"    [OK] {name}")
            print("")
            print(f"  Permanent Sheet URL (always the same):")
            print(f"  {permanent_url}")
            print("")
            print("="*60)
            print("")

            return permanent_url

        except Exception as e:
            print(f"\n[ERROR] Failed importing to permanent sheet: {e}")
            print(f"Full error: {traceback.format_exc()}")
            return None

    def upload_to_google_sheets(self, main_df, feedback_df=None, original_df=None, excel_file_path=None):
        """
        Upload Excel file to Google Drive using user's Google account (OAuth)
        Args: main_df - main data, feedback_df - feedback data, original_df - original CSV data,
              excel_file_path - path to Excel file (contains all sheets including original data)
        Returns: Google Sheet URL or None
        """
        try:
            print("\n" + "="*60)
            print("           UPLOADING TO GOOGLE DRIVE")
            print("="*60)

            # Get OAuth credentials
            print("Authenticating with your Google account...")
            credentials = self.get_oauth_credentials()

            if not credentials:
                return None

            # Build Drive service for file upload
            drive_service = build('drive', 'v3', credentials=credentials)

            # Your Google Drive folder ID
            folder_id = "1QsBV9mV3ATrZ6qU-QUIi1hBRlQ2pBk0c"

            uploaded_file_url = None

            # Upload and convert Excel to Google Sheets (preserves charts!)
            if excel_file_path and os.path.exists(excel_file_path):
                print("\n[STEP 1] Converting Excel to Google Sheets (with charts)...")

                # Generate unique name with timestamp
                base_name = os.path.splitext(os.path.basename(excel_file_path))[0]
                new_name = f"{base_name}_GoogleSheets"

                file_metadata = {
                    'name': new_name,
                    'mimeType': 'application/vnd.google-apps.spreadsheet',  # Convert to Google Sheets
                    'parents': [folder_id]
                }

                media = MediaFileUpload(
                    excel_file_path,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    resumable=True
                )

                # Upload and convert
                uploaded_file = drive_service.files().create(
                    body=file_metadata,
                    media_body=media,
                    fields='id, name, webViewLink'
                ).execute()

                print(f"   [OK] Converted: {uploaded_file.get('name')}")
                print(f"   [OK] All sheets preserved (including charts!)")
                uploaded_file_url = uploaded_file.get('webViewLink')
                temp_spreadsheet_id = uploaded_file.get('id')

                # Import to permanent Google Sheet
                print(f"\n[STEP 2] Importing to permanent Google Sheet...")
                permanent_url = self.import_to_permanent_sheet(temp_spreadsheet_id, credentials)

                if permanent_url:
                    return permanent_url
                else:
                    # If import failed, return temp file URL as fallback
                    print("[WARNING] Import failed, returning temporary file URL")
                    return uploaded_file_url

            return None

        except Exception as e:
            print(f"\n[ERROR] Failed uploading to Google Sheets: {e}")
            print(f"Full error: {traceback.format_exc()}")
            return None

def main():
    """
    Main execution function
    """
    print("="*60)
    print("    TICKETS TO GPRS FORMATTER - FINAL VERSION")
    print("="*60)
    
    try:
        # Initialize formatter
        formatter = TicketsToGPRSFormatter()
        
        # Load latest tickets data
        tickets_df, source_file = formatter.load_latest_tickets()
        
        # Transform to GPRS format
        gprs_df = formatter.transform_to_gprs_format(tickets_df)
        
        # Create feedback sheet and remove entries from main sheet
        feedback_df, gprs_df = formatter.create_feedback_sheet(tickets_df, gprs_df)
        
        # Create daily ticket analysis pivot table
        pivot_data = formatter.create_daily_ticket_analysis(gprs_df)

        # Save Excel file (including original data)
        output_path = formatter.save_excel_file(gprs_df, feedback_df, pivot_data, tickets_df)
        
        # Generate summary
        formatter.generate_summary(gprs_df, feedback_df)
        
        # Success message
        source_filename = os.path.basename(source_file)
        output_filename = os.path.basename(output_path)
        
        print(f"\nSUCCESS!")
        print(f"Converted: {source_filename}")
        print(f"Output: {output_filename}")
        print(f"Location: {formatter.transformed_data_folder}")

        # Ask user if they want to upload to Google Drive
        print("\n" + "="*60)
        user_response = input("Upload to Google Drive? (yes/no): ").strip().lower()

        if user_response in ['yes', 'y']:
            sheet_url = formatter.upload_to_google_sheets(
                gprs_df,
                feedback_df,
                tickets_df,
                excel_file_path=output_path
            )
            if sheet_url:
                print(f"\n[SUCCESS] All done! Your file is now in Google Drive!")
                print(f"[INFO] Excel file includes original data as 'original Data' sheet")
                print(f"[LINK] View here: {sheet_url}")
        else:
            print("\nSkipped Google Drive upload. Data saved locally only.")

        return True
        
    except Exception as e:
        print(f"\nERROR: {e}")
        print(f"Full error details:")
        print(traceback.format_exc())
        return False

if __name__ == "__main__":
    success = main()
    if not success:
        sys.exit(1)